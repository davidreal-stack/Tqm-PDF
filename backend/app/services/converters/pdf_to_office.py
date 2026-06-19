"""
PDF → Office / Texto / HTML
"""
from __future__ import annotations

import asyncio
from pathlib import Path

import pdfplumber
from pdf2docx import Converter as Pdf2DocxConverter

from app.core.logging import logger


class PdfToOfficeConverter:

    async def to_docx(self, input_path: Path, output_dir: Path, **kwargs) -> Path:
        """PDF → DOCX con pdf2docx (preserva layout)."""
        return await asyncio.get_event_loop().run_in_executor(
            None, self._to_docx_sync, input_path, output_dir
        )

    async def to_xlsx(self, input_path: Path, output_dir: Path) -> Path:
        """PDF → XLSX extrayendo tablas con pdfplumber."""
        return await asyncio.get_event_loop().run_in_executor(
            None, self._to_xlsx_sync, input_path, output_dir
        )

    async def to_txt(self, input_path: Path, output_dir: Path) -> Path:
        """PDF → TXT plano."""
        return await asyncio.get_event_loop().run_in_executor(
            None, self._to_txt_sync, input_path, output_dir
        )

    async def to_html(self, input_path: Path, output_dir: Path) -> Path:
        """PDF → HTML básico."""
        return await asyncio.get_event_loop().run_in_executor(
            None, self._to_html_sync, input_path, output_dir
        )

    # ── DOCX ─────────────────────────────────────────────────────────────────

    def _to_docx_sync(self, input_path: Path, output_dir: Path) -> Path:
        output_path = output_dir / f"{input_path.stem}.docx"
        log = logger.bind(input=input_path.name)

        cv = Pdf2DocxConverter(str(input_path))
        cv.convert(str(output_path), start=0, end=None)
        cv.close()

        log.info("PDF → DOCX completado", output=output_path.name)
        return output_path

    # ── XLSX ─────────────────────────────────────────────────────────────────

    def _to_xlsx_sync(self, input_path: Path, output_dir: Path) -> Path:
        import openpyxl

        output_path = output_dir / f"{input_path.stem}.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # quitar hoja vacía por defecto

        with pdfplumber.open(str(input_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                page_text = page.extract_text() or ""

                if tables:
                    for t_idx, table in enumerate(tables):
                        sheet_name = f"P{page_num}_T{t_idx+1}"[:31]
                        ws = wb.create_sheet(title=sheet_name)
                        for row in table:
                            ws.append([cell or "" for cell in row])
                else:
                    # Sin tablas → volcamos el texto crudo en una hoja de texto
                    sheet_name = f"Página {page_num}"[:31]
                    ws = wb.create_sheet(title=sheet_name)
                    for line in page_text.split("\n"):
                        ws.append([line])

        if not wb.worksheets:
            wb.create_sheet("Sin contenido")

        wb.save(output_path)
        logger.info("PDF → XLSX completado", output=output_path.name)
        return output_path

    # ── TXT ──────────────────────────────────────────────────────────────────

    def _to_txt_sync(self, input_path: Path, output_dir: Path) -> Path:
        output_path = output_dir / f"{input_path.stem}.txt"

        with pdfplumber.open(str(input_path)) as pdf:
            lines = []
            for i, page in enumerate(pdf.pages, start=1):
                lines.append(f"{'='*60}")
                lines.append(f"  Página {i}")
                lines.append(f"{'='*60}")
                text = page.extract_text() or "(página sin texto — puede ser imagen escaneada)"
                lines.append(text)
                lines.append("")

        output_path.write_text("\n".join(lines), encoding="utf-8")
        logger.info("PDF → TXT completado", output=output_path.name)
        return output_path

    # ── HTML ─────────────────────────────────────────────────────────────────

    def _to_html_sync(self, input_path: Path, output_dir: Path) -> Path:
        output_path = output_dir / f"{input_path.stem}.html"

        with pdfplumber.open(str(input_path)) as pdf:
            total = len(pdf.pages)
            pages_html = []

            for i, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                paragraphs = "".join(
                    f"<p>{line}</p>" for line in text.split("\n") if line.strip()
                )
                pages_html.append(
                    f'<section class="page" id="page-{i}">'
                    f'<div class="page-header">Página {i} de {total}</div>'
                    f'{paragraphs}'
                    f'</section>'
                )

        html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>{input_path.stem}</title>
<style>
  body {{ font-family: Georgia, serif; max-width: 820px; margin: 2rem auto; padding: 0 1rem; color: #222; line-height: 1.7; }}
  .page {{ border-bottom: 1px solid #ddd; padding: 2rem 0; }}
  .page-header {{ font-size: 0.75rem; color: #888; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 1rem; }}
  p {{ margin: 0.5rem 0; }}
</style>
</head>
<body>
{''.join(pages_html)}
</body>
</html>"""

        output_path.write_text(html, encoding="utf-8")
        logger.info("PDF → HTML completado", output=output_path.name)
        return output_path